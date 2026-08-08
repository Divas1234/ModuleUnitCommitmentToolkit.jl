import openpyxl
import random
import math

def generate_stochastic_load_curve(n_days=7, base_load=280.0, volatility=25.0):
    random.seed(42)
    load_values = []
    current_load = base_load

    for day in range(1, n_days + 1):
        for h in range(1, 25):
            diurnal_component = 30.0 * math.sin(2 * math.pi * (h - 6) / 24)
            stochastic_noise = random.gauss(0, 1) * volatility * 0.4
            drift = 0.15 * (base_load - current_load) + random.gauss(0, 1) * 8.0
            current_load = current_load + drift

            val = base_load + diurnal_component + stochastic_noise + (current_load - base_load) * 0.3
            val = max(180.0, min(420.0, val))
            load_values.append(round(val, 2))

    return load_values

def main():
    excel_path = "data/data.xlsx"
    print(f"Updating load_curve in {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    if "load_curve" not in wb.sheetnames:
        raise ValueError("Sheet 'load_curve' not found in data.xlsx")
    
    ws = wb["load_curve"]
    load_vals = generate_stochastic_load_curve(n_days=7)
    
    for i, val in enumerate(load_vals, start=1):
        ws.cell(row=i+1, column=1, value=float(i))
        ws.cell(row=i+1, column=2, value=float(val))
    
    wb.save(excel_path)
    print(f"✓ Successfully updated 168-hour stochastic load curve in {excel_path}!")

if __name__ == "__main__":
    main()

